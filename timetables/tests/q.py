from openpyxl import load_workbook
import json
import os
import hashlib
import time
import re

week = 0
def make_json(name_file,massiv):
	with open(name_file, 'w') as fw:
	    json.dump(massiv, fw)
	return os.path.getsize(name_file) / 1000000
def out_red(text):
    print("\033[31m {}" .format(text))
def out_yellow(text):
    print("\033[33m {}" .format(text))
def out_blue(text):
    print("\033[34m {}" .format(text))

out_red("Python 3.0 addon for vk bot 'Рапсписание КБСП' for online parsing http://mirea.ru")
time.sleep(1)
print("Разработка: @Miron_root")
time.sleep(1)

out_red("Starting")

def load_json(name_file):
        with open(name_file, 'r') as fr:
                return(json.load(fr))
def md5summ(filename):
	return hashlib.md5(open(filename, 'rb').read()).hexdigest()

out_red("Загрузка расписания с офф сайта МИРЭА")
print("")
os.system("rm table.xlsx")

os.system("wget -O table.xlsx https://webservices.mirea.ru/upload/iblock/390/%D0%9A%D0%91%D0%B8%D0%A1%D0%9F%201%20%D0%BA%D1%83%D1%80%D1%81%202%20%D1%81%D0%B5%D0%BC-%D0%94.xlsx")
out_red("Завершено!")

wb = load_workbook('table.xlsx')
sheet = wb['Лист1']
print("--------------------------------")
# = load_json("prefix.json")
def colnum_string(n):
    str = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        str = chr(65 + remainder) + str
    return str
result = []
mass3 = []
out_yellow("парсинг Расписания")
for i in range(1,sheet.max_column):
	if(str(sheet[colnum_string(i)+"2"].value) == "None"):
			o=1
	else:
		try:
			mass3.append({"group":re.findall("(\w+\w+\w+\w+[-]\d\d[-]\d\d)", str(sheet[colnum_string(i)+"2"].value))[0], "para":colnum_string(i),"type":colnum_string(i+1),"fio":colnum_string(i+2),"aud":colnum_string(i+3)})
		except IndexError:
			o=0
mass_prefix = []
out_yellow("создание prefix.json")

for i in range(len(mass3)):
			mass_prefix.append(mass3[i]["group"])
make_json("prefix.json",mass_prefix)
#print(mass3)
#print(len(mass3))
if(len(mass_prefix) != len(mass3)):
		exit
else:
	out_yellow("Создано! длинна: " + str(len(mass_prefix)))

q = 0
w = 0
num_par =0
day = 1
mass1 = [[],[],[],[],[],[],[]]

def makeJSON_RASP(week,mass,prefix):
		q = 0
		w = 0
		num_par =0
		day = 1
		mass1 = [[],[],[],[],[],[],[]]
		column_lesson = mass[0]                 #название пары                  ББСО-04-20
		colum_type = mass[1]                    #тип пары практик/лекция
		colum_fio = mass[2]                     #фио препода
		colum_aud =  mass[3]  					#аудитория
		out_blue("Создание расписание для группы:"+str(prefix))
		for i in range(4,76):
			w +=1
			num_par = sheet['FZ'+str(i-week)].value #это не трогать работает
			if q == 12:
				q = 0
#				print("-----------------------------------")
				day = day + 1
			if i % 2 == week:

				if(str(sheet[column_lesson+str(i)].value) == "None"):
					o=1
				else:
#					print(num_par, "   ", str(sheet[column_lesson+str(i)].value).replace("\n", " "), "     ", sheet[colum_type+str(i)].value, "     ", str(sheet[colum_fio+str(i)].value).replace("\n", " "), "   ", str(sheet[colum_aud+str(i)].value).replace("\n", " "))
					buffer = {}
					buffer["num_para"] = str(num_par)
					buffer["para"] = str(sheet[column_lesson+str(i)].value).replace("\n", " ")
					buffer["type"] = str(sheet[colum_type+str(i)].value).replace("\n", " ")
					buffer["fio"]  = str(sheet[colum_fio+str(i)].value).replace("\n", " ")
					buffer["aud"]  = str(sheet[colum_aud+str(i)].value).replace("\n", " ")
					mass1[day].append(buffer)
			q +=1
#		filemd5 = md5summ(prefix+"/nechet.json")
		if (week == 0):
			os.system("sudo mkdir " + prefix)
			os.system("sudo chmod 777 "+prefix + "/")
			make_json(prefix+"/nechet.json",mass1)
			out_blue("Сохранено нечетное расписание для группы " + str(prefix) + " по пути " + str(prefix) + "/nechet.json")
			print("=======================================")
		else:
			os.system("sudo mkdir " + prefix)
			os.system("sudo chmod 777 "+prefix + "/")
			make_json(prefix+"/chet.json",mass1)
			out_blue("Сохранено четное расписание для группы " + str(prefix) + " по пути " + str(prefix) + "/chet.json")
			print("=======================================")
#		if (md5summ(prefix+"/chet.json") == filemd5):
#					out_yellow("Расписание группы: "+prefix+" не изменено")
#		else:
#					out_red("Расписание группы: "+prefix+" изменилось")
#					print(filemd5 + " " + md5summ(prefix+"/chet.json"))
		q = 0
		w = 0
		num_par =0
		mass1 = [[],[],[],[],[],[],[]]
		day = 1

for week in range(2):
	for i in range(len(mass3)):
		mass = [mass3[i]["para"],mass3[i]["type"],mass3[i]["fio"],mass3[i]["aud"]]
		makeJSON_RASP(week,mass,mass3[i]["group"])
	i = 0
